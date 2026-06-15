#!/usr/bin/env python3
"""Append metadata rows to labels.xlsx from audio files (iTunes / local tags).

Reads embedded tags from files in audio/ (.m4a, .mp3, .wav, .flac).
Fills metadata columns only; leaves sa_note and verified blank for you.

Usage:
    python3 scripts/fill_labels_from_audio.py
    python3 scripts/fill_labels_from_audio.py --dry-run
    python3 scripts/fill_labels_from_audio.py --file audio/my_song.m4a

Then sync to CSV:
    python3 scripts/sync_labels_from_excel.py
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing openpyxl. Install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from mutagen import File as MutagenFile
except ImportError:
    print("Missing mutagen. Install: pip3 install mutagen", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
AUDIO_DIR = ROOT / "audio"
XLSX = ROOT / "labels.xlsx"
SHEET = "labels"
DATA_START_ROW = 3

AUDIO_EXTENSIONS = {".m4a", ".mp3", ".wav", ".flac", ".aac", ".MP4", ".M4A", ".MP3"}

COLUMNS = [
    "track_id",
    "title",
    "film",
    "year",
    "singer",
    "artists",
    "music_director",
    "writers",
    "genre",
    "label",
    "isrc",
    "track_no",
    "audio_filename",
    "sa_note",
    "verified",
    "notes",
]

# Manual columns the user fills; never auto-populated.
MANUAL_COLUMNS = {"sa_note", "verified", "notes"}

COLUMN_HINTS = {
    "track_id": "Unique ID (auto)",
    "title": "Song title (auto)",
    "film": "Film / album (auto)",
    "year": "Release year (auto)",
    "singer": "Primary singer (auto)",
    "artists": "All credited artists (auto)",
    "music_director": "Music director / composer (auto)",
    "writers": "Composer & lyricist credit (auto)",
    "genre": "Genre (auto)",
    "label": "Record label (auto)",
    "isrc": "ISRC code (auto)",
    "track_no": "Track number (auto)",
    "audio_filename": "File name inside audio/ (auto)",
    "sa_note": "Sa pitch: C, C#, ... B (you fill)",
    "verified": "true when verified on keyboard (you fill)",
    "notes": "Optional notes",
}


def slug(text: str, max_len: int = 40) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return text[:max_len].strip("-") or "track"


def parse_label(cprt: str | None) -> str:
    """'℗ 2017 Saregama' -> 'Saregama'."""
    if not cprt:
        return ""
    s = str(cprt).strip()
    s = re.sub(r"^[℗©Ⓟⓒ()CcPp\s]+", "", s)   # leading copyright marks
    s = re.sub(r"^(19|20)\d{2}\s*", "", s)      # leading year
    return s.strip()


def parse_isrc(xid: str | None) -> str:
    """'Orchard:isrc:INH100161270' -> 'INH100161270'."""
    if not xid:
        return ""
    s = str(xid)
    if "isrc:" in s.lower():
        idx = s.lower().rindex("isrc:") + len("isrc:")
        return s[idx:].strip()
    return ""


def _first(value) -> str:
    if not value:
        return ""
    if isinstance(value, (list, tuple)):
        return str(value[0]) if value else ""
    return str(value)


def _atom(tags, key: str) -> str:
    try:
        return str(tags.get(key, [""])[0] or "")
    except Exception:
        return ""


def parse_year(value: str | None) -> str:
    if not value:
        return ""
    match = re.search(r"(19|20)\d{2}", str(value))
    return match.group(0) if match else ""


def first_artist(artist: str | None) -> str:
    if not artist:
        return ""
    # iTunes / ID3: "Artist A & Artist B" or "Artist A, Artist B"
    for sep in ("/", ",", ";", "&", " feat.", " ft.", " featuring "):
        if sep in artist:
            artist = artist.split(sep)[0]
    return artist.strip()


def extract_facets(path: Path) -> dict[str, str]:
    """Pull every auto-derivable metadata facet from an audio file's tags."""
    out = {
        "title": "", "film": "", "year": "", "singer": "", "artists": "",
        "music_director": "", "writers": "", "genre": "", "label": "",
        "isrc": "", "track_no": "",
    }

    easy = MutagenFile(path, easy=True)
    if easy is not None and easy.tags:
        t = easy.tags
        out["title"] = _first(t.get("title"))
        out["film"] = _first(t.get("album"))
        out["artists"] = _first(t.get("artist"))
        out["music_director"] = _first(t.get("albumartist"))
        out["writers"] = _first(t.get("composer"))
        out["genre"] = _first(t.get("genre"))
        out["year"] = parse_year(_first(t.get("date")))
        out["isrc"] = _first(t.get("isrc"))
        tn = _first(t.get("tracknumber"))
        if tn:
            out["track_no"] = tn.split("/")[0].strip()

    # Raw MP4 atoms for iTunes purchases (carry label, ISRC, etc.)
    if path.suffix.lower() in (".m4a", ".mp4", ".aac"):
        raw = MutagenFile(path)
        if raw is not None and getattr(raw, "tags", None):
            a = raw.tags
            out["title"] = out["title"] or _atom(a, "\xa9nam")
            out["film"] = out["film"] or _atom(a, "\xa9alb")
            out["artists"] = out["artists"] or _atom(a, "\xa9ART")
            out["music_director"] = out["music_director"] or _atom(a, "aART")
            out["writers"] = out["writers"] or _atom(a, "\xa9wrt")
            out["genre"] = out["genre"] or _atom(a, "\xa9gen")
            if not out["year"]:
                out["year"] = parse_year(_atom(a, "\xa9day"))
            if not out["label"]:
                out["label"] = parse_label(_atom(a, "cprt"))
            if not out["isrc"]:
                out["isrc"] = parse_isrc(_atom(a, "xid "))
            if not out["track_no"]:
                trkn = a.get("trkn")
                if trkn:
                    try:
                        out["track_no"] = str(trkn[0][0])
                    except Exception:
                        pass

    out["singer"] = first_artist(out["artists"])
    return out


def read_tags(path: Path) -> dict[str, str]:
    f = extract_facets(path)

    title = f["title"] or path.stem.replace("_", " ")
    singer = f["singer"]
    year = f["year"]

    track_id = slug(title)
    if singer:
        track_id = f"{track_id}-{slug(singer, 24)}"
    if year:
        track_id = f"{track_id}-{year}"

    row = {key: "" for key in COLUMNS}
    row.update(f)
    row["track_id"] = track_id
    row["title"] = title
    row["audio_filename"] = path.name
    return row


def load_existing_filenames() -> set[str]:
    if not XLSX.exists():
        return set()
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb[SHEET]
    header = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "audio_filename" not in header:
        wb.close()
        return set()
    col = header.index("audio_filename")
    names: set[str] = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row and len(row) > col and row[col]:
            names.add(str(row[col]).strip())
    wb.close()
    return names


def collect_audio_paths(single: Path | None) -> list[Path]:
    if single:
        if not single.is_file():
            raise FileNotFoundError(single)
        return [single]
    if not AUDIO_DIR.is_dir():
        return []
    files = [p for p in sorted(AUDIO_DIR.iterdir()) if p.is_file() and p.suffix in AUDIO_EXTENSIONS]
    return [p for p in files if not p.name.startswith(".")]


def append_rows(rows: list[dict[str, str]], dry_run: bool) -> int:
    if not rows:
        return 0
    if dry_run:
        for r in rows:
            print(f"[dry-run] would add: {r['audio_filename']} → {r['track_id']} | {r['title']} | {r['singer']}")
        return len(rows)

    if not XLSX.exists():
        print(f"Not found: {XLSX}", file=sys.stderr)
        return 0

    wb = load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        print(f"Sheet '{SHEET}' missing in {XLSX}", file=sys.stderr)
        return 0

    ws = wb[SHEET]
    next_row = ws.max_row + 1
    if next_row < DATA_START_ROW:
        next_row = DATA_START_ROW

    for r in rows:
        for col, key in enumerate(COLUMNS, start=1):
            ws.cell(row=next_row, column=col, value=r[key])
        next_row += 1

    wb.save(XLSX)
    wb.close()
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill labels.xlsx metadata from audio file tags.")
    parser.add_argument("--file", type=Path, help="Single audio file (default: scan audio/)")
    parser.add_argument("--dry-run", action="store_true", help="Print rows without writing Excel")
    args = parser.parse_args()

    try:
        paths = collect_audio_paths(args.file)
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 1

    if not paths:
        print(f"No audio files found in {AUDIO_DIR} (supported: m4a, mp3, wav, flac, aac)")
        return 0

    existing = load_existing_filenames()
    new_rows: list[dict[str, str]] = []

    for path in paths:
        if path.name in existing:
            print(f"Skip (already in Excel): {path.name}")
            continue
        meta = read_tags(path)
        new_rows.append(meta)
        if not args.dry_run:
            print(f"Add: {path.name} → {meta['title']} | {meta['singer']} | film/album: {meta['film']}")

    count = append_rows(new_rows, args.dry_run)
    if count and not args.dry_run:
        print(f"Added {count} row(s) to {XLSX}")
        print("Next: fill sa_note + verified, then run: python3 scripts/sync_labels_from_excel.py")
    elif args.dry_run:
        print(f"Would add {count} row(s)")
    else:
        print("Nothing new to add.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
