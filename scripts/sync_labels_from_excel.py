#!/usr/bin/env python3
"""Copy label rows from labels.xlsx into labels.csv.

Edit labels.xlsx in Excel/Numbers/Google Sheets (export as xlsx), then run:

    python3 scripts/sync_labels_from_excel.py

Or ask the agent: "sync labels from excel to csv".
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing openpyxl. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "labels.xlsx"
CSV = ROOT / "labels.csv"
SHEET = "labels"

COLUMNS = [
    "track_id",
    "title",
    "film",
    "year",
    "singer",
    "artists",
    "music_director",
    "writers",
    "genre",
    "label",
    "isrc",
    "track_no",
    "audio_filename",
    "sa_note",
    "verified",
    "notes",
]

# Metadata facet columns added automatically; tolerated if missing in older sheets.
OPTIONAL_COLUMNS = {
    "artists", "music_director", "writers", "genre", "label", "isrc", "track_no",
}

CSV_FIELDS = [
    "track_id",
    "audio_path",
    "sa_note",
    "verified",
    "title",
    "film",
    "year",
    "singer",
    "artists",
    "music_director",
    "writers",
    "genre",
    "label",
    "isrc",
    "track_no",
    "notes",
]


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main() -> int:
    if not XLSX.exists():
        print(f"Not found: {XLSX}", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"Sheet '{SHEET}' not found. Sheets: {wb.sheetnames}", file=sys.stderr)
        return 1

    ws = wb[SHEET]
    header = [cell_str(c).lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = [c for c in COLUMNS if c not in OPTIONAL_COLUMNS]
    missing = [c for c in required if c not in header]
    if missing:
        wb.close()
        print(f"Missing columns in row 1: {missing}", file=sys.stderr)
        print(f"Expected at least: {required}", file=sys.stderr)
        return 1

    idx = {name: header.index(name) for name in COLUMNS if name in header}
    out_rows: list[dict[str, str]] = []

    # Row 1 = headers, row 2 = hint text, row 3+ = data
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        data = {
            col: (cell_str(row[idx[col]]) if col in idx and idx[col] < len(row) else "")
            for col in COLUMNS
        }
        if not data["track_id"]:
            print(f"Skipping row {row_num}: empty track_id", file=sys.stderr)
            continue
        out_rows.append(data)

    wb.close()

    with CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for r in out_rows:
            filename = r["audio_filename"]
            audio_path = f"audio/{filename}" if filename else ""
            verified = r["verified"].lower() in ("true", "yes", "1", "y")
            writer.writerow(
                {
                    "track_id": r["track_id"],
                    "audio_path": audio_path,
                    "sa_note": r["sa_note"],
                    "verified": "true" if verified else "false",
                    "title": r["title"],
                    "film": r["film"],
                    "year": r["year"],
                    "singer": r["singer"],
                    "artists": r.get("artists", ""),
                    "music_director": r.get("music_director", ""),
                    "writers": r.get("writers", ""),
                    "genre": r.get("genre", ""),
                    "label": r.get("label", ""),
                    "isrc": r.get("isrc", ""),
                    "track_no": r.get("track_no", ""),
                    "notes": r["notes"],
                }
            )

    print(f"Wrote {len(out_rows)} row(s) to {CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
