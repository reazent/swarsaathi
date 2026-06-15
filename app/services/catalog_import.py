from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.db.models import PitchResult, Track

SHEET = "labels"
DATA_START_ROW = 3
COLUMNS = [
    "track_id",
    "title",
    "film",
    "year",
    "singer",
    "artists",
    "music_director",
    "writers",
    "genre",
    "label",
    "isrc",
    "track_no",
    "audio_filename",
    "sa_note",
    "verified",
    "notes",
]


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _search_blob(*parts: str) -> str:
    return " ".join(p.lower() for p in parts if p)


def load_rows_from_xlsx(path: Path | None = None) -> list[dict[str, str]]:
    xlsx = path or settings.labels_xlsx
    if not xlsx.exists():
        return []

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET]
    header = [_cell_str(c).lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: header.index(name) for name in COLUMNS if name in header}

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        data = {col: _cell_str(row[idx[col]]) if col in idx and idx[col] < len(row) else "" for col in COLUMNS}
        if not data.get("track_id"):
            continue
        rows.append(data)
    wb.close()
    return rows


def import_catalog(db: Session) -> tuple[int, int]:
    """Upsert tracks from labels.xlsx; skip rows whose audio file is missing."""
    imported = 0
    skipped = 0

    for row in load_rows_from_xlsx():
        filename = row.get("audio_filename", "")
        if not filename:
            skipped += 1
            continue
        audio_file = settings.audio_dir / filename
        if not audio_file.is_file():
            skipped += 1
            continue

        track_id = row["track_id"]
        title = row.get("title", "")
        film = row.get("film", "")
        year = row.get("year", "")
        singer = row.get("singer", "")
        artists = row.get("artists", "")
        music_director = row.get("music_director", "")
        writers = row.get("writers", "")
        genre = row.get("genre", "")
        label = row.get("label", "")
        isrc = row.get("isrc", "")
        track_no = row.get("track_no", "")

        track = db.get(Track, track_id)
        if track is None:
            track = Track(track_id=track_id)
            db.add(track)
            imported += 1

        track.title = title
        track.film = film
        track.year = year
        track.singer = singer
        track.artists = artists
        track.music_director = music_director
        track.writers = writers
        track.genre = genre
        track.label = label
        track.isrc = isrc or None
        track.track_no = track_no
        track.audio_filename = filename
        track.audio_path = f"audio/{filename}"
        track.search_blob = _search_blob(
            title, film, singer, artists, music_director, writers, genre, label, year
        )

        sa = row.get("sa_note", "").strip()
        verified = row.get("verified", "").lower() in ("true", "yes", "1", "y")
        if sa:
            pitch = db.get(PitchResult, track_id)
            if pitch is None:
                pitch = PitchResult(track_id=track_id, sa_note=sa, confidence=1.0, analysis_version="human-label")
                db.add(pitch)
            pitch.label_sa_note = sa
            pitch.label_verified = verified
            if verified:
                pitch.sa_note = sa
                pitch.confidence = 1.0

    db.commit()
    return imported, skipped
